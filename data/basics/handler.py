import csv
import json
from openpyxl import Workbook, load_workbook
from typing import List, Dict, Any, Type
import cx_Oracle

class CSVHandler:
    "CSV fájlok írása és olvasása"
    
    @staticmethod
    def write(data: List[Any], filename: str, fields: List[str]):
        "CSV írás fejléccel"
        with open(filename, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=fields, delimiter=';')
            writer.writeheader()
            for item in data:
                writer.writerow(item.to_dict())
    
    @staticmethod
    def read(filename: str, data_type: Type) -> List[Any]:
        "CSV olvasás fejléccel"
        with open(filename, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f, delimiter=';')
            return [data_type(**row) for row in reader]


class JSONHandler:
    "JSON fájlok írása és olvasása"
    
    @staticmethod
    def write(data: List[Any], filename: str):
        "JSON írás lista formátumban"
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump([item.to_dict() for item in data], f, 
                     indent=2, ensure_ascii=False)
    
    @staticmethod
    def read(filename: str, data_type: Type) -> List[Any]:
        "JSON olvasás"
        with open(filename, 'r', encoding='utf-8') as f:
            data = json.load(f)
            return [data_type(**item) for item in data]


class XLSXHandler:
    "XLSX fájlok írása és olvasása - minden típus külön munkalapon"
    
    @staticmethod
    def write(data_dict: Dict[str, tuple], filename: str):
        wb = Workbook()
        wb.remove(wb.active)  # Alapértelmezett lap törlése
        
        for sheet_name, (data, fields) in data_dict.items():
            ws = wb.create_sheet(title=sheet_name)
            
            # Fejléc írása
            ws.append(fields)
            
            # Adatok írása
            for item in data:
                row_data = [item.to_dict()[field] for field in fields]
                ws.append(row_data)
        
        wb.save(filename)
    
    @staticmethod
    def read(filename: str, sheet_names: List[str], 
             data_types: Dict[str, Type]) -> Dict[str, List[Any]]:
        wb = load_workbook(filename)
        result = {}
        
        for sheet_name in sheet_names:
            ws = wb[sheet_name]
            headers = [cell.value for cell in ws[1]]
            
            data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers, row))
                data.append(data_types[sheet_name](**row_dict))
            
            result[sheet_name] = data
        

        return result
