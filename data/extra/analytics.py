# data/extra/analytics.py
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import ColorScaleRule
from typing import List, Dict
from collections import Counter

class DataAnalytics:
    
    def __init__(self, students: List[any], universities: List[any], 
                 courses: List[any], enrollments: List[any]):
        self.students = students
        self.universities = universities
        self.courses = courses
        self.enrollments = enrollments
    
    def generate_analytics_report(self, filename: str):
        """Elemzési jelentés generálása diagramokkal"""
        wb = Workbook()
        
        # 1. Hallgatók egyetemenkénti eloszlása (Bar Chart)
        self._create_university_distribution(wb)
        
        # 2. Szakok népszerűsége (Pie Chart)
        self._create_major_distribution(wb)
        
        # 3. GPA eloszlás (Oszlopdiagram feltételes formázással)
        self._create_gpa_analysis(wb)
        
        # 4. Beiratkozási trendek (Line Chart)
        self._create_enrollment_trends(wb)
        
        # Alapértelmezett lap törlése
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        wb.save(filename)
        print(f"Elemzési jelentés létrehozva: {filename}")
    
    def _create_university_distribution(self, wb: Workbook):
        """Hallgatók egyetemenkénti eloszlása oszlopdiagrammal"""
        ws = wb.create_sheet("Egyetem Eloszlás")
        
        # Adatok számítása
        uni_counts = Counter(s.university_id for s in self.students)
        uni_dict = {u.university_id: u.name for u in self.universities}
        
        # Fejléc
        ws['A1'] = 'Egyetem'
        ws['B1'] = 'Hallgatók száma'
        ws['A1'].font = Font(bold=True)
        ws['B1'].font = Font(bold=True)
        
        # Adatok írása
        row = 2
        for uni_id, count in sorted(uni_counts.items()):
            ws[f'A{row}'] = uni_dict.get(uni_id, f'University {uni_id}')
            ws[f'B{row}'] = count
            row += 1
        
        # Oszlopdiagram létrehozása[2][23]
        chart = BarChart()
        chart.title = "Hallgatók Eloszlása Egyetemek Szerint"
        chart.x_axis.title = "Egyetem"
        chart.y_axis.title = "Hallgatók Száma"
        
        data = Reference(ws, min_col=2, min_row=1, max_row=row-1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=row-1)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 15
        chart.width = 25
        
        ws.add_chart(chart, "D2")
    
    def _create_major_distribution(self, wb: Workbook):
        """Szakok megoszlása kördiagrammal"""
        ws = wb.create_sheet("Szak Eloszlás")
        
        # Adatok számítása
        major_counts = Counter(s.major for s in self.students)
        
        # Fejléc
        ws['A1'] = 'Szak'
        ws['B1'] = 'Hallgatók száma'
        ws['A1'].font = Font(bold=True)
        ws['B1'].font = Font(bold=True)
        
        # Adatok írása
        row = 2
        for major, count in sorted(major_counts.items(), 
                                   key=lambda x: x[1], reverse=True):
            ws[f'A{row}'] = major
            ws[f'B{row}'] = count
            row += 1
        
        # Kördiagram létrehozása
        chart = PieChart()
        chart.title = "Szakok Megoszlása"
        
        data = Reference(ws, min_col=2, min_row=1, max_row=row-1)
        labels = Reference(ws, min_col=1, min_row=2, max_row=row-1)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        chart.height = 15
        chart.width = 20
        
        ws.add_chart(chart, "D2")
    
    def _create_gpa_analysis(self, wb: Workbook):
        """GPA elemzés feltételes formázással"""
        ws = wb.create_sheet("GPA Elemzés")
        
        # Fejléc
        headers = ['Hallgató ID', 'Név', 'Egyetem', 'Szak', 'GPA']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", 
                                   end_color="366092", 
                                   fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
        
        # Adatok írása
        uni_dict = {u.university_id: u.name for u in self.universities}
        for row, student in enumerate(self.students, 2):
            ws.cell(row=row, column=1, value=student.student_id)
            ws.cell(row=row, column=2, 
                   value=f"{student.first_name} {student.last_name}")
            ws.cell(row=row, column=3, 
                   value=uni_dict.get(student.university_id, 'N/A'))
            ws.cell(row=row, column=4, value=student.major)
            ws.cell(row=row, column=5, value=student.gpa)
        
        # Feltételes formázás a GPA oszlopra
        gpa_range = f'E2:E{len(self.students) + 1}'
        
        # Színskála: piros (2.0) -> sárga (3.0) -> zöld (4.0)
        color_scale = ColorScaleRule(
            start_type='num', start_value=2.0, start_color='F8696B',
            mid_type='num', mid_value=3.0, mid_color='FFEB84',
            end_type='num', end_value=4.0, end_color='63BE7B'
        )
        ws.conditional_formatting.add(gpa_range, color_scale)
        
        # Oszlopszélességek
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 10
    
    def _create_enrollment_trends(self, wb: Workbook):
        """Beiratkozási trendek vonaldiagrammal"""
        ws = wb.create_sheet("Beiratkozási Trendek")
        
        # Beiratkozások évenkénti összesítése
        from datetime import datetime
        year_counts = Counter()
        for enr in self.enrollments:
            year = datetime.strptime(enr.enrollment_date, '%Y-%m-%d').year
            year_counts[year] += 1
        
        # Fejléc
        ws['A1'] = 'Év'
        ws['B1'] = 'Beiratkozások'
        ws['A1'].font = Font(bold=True)
        ws['B1'].font = Font(bold=True)
        
        # Adatok írása
        row = 2
        for year in sorted(year_counts.keys()):
            ws[f'A{row}'] = year
            ws[f'B{row}'] = year_counts[year]
            row += 1
        
        # Vonaldiagram létrehozása
        chart = LineChart()
        chart.title = "Beiratkozási Trend Időben"
        chart.x_axis.title = "Év"
        chart.y_axis.title = "Beiratkozások Száma"
        
        data = Reference(ws, min_col=2, min_row=1, max_row=row-1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=row-1)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 15
        chart.width = 25
        
        ws.add_chart(chart, "D2")
